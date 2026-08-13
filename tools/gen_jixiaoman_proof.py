# -*- coding: utf-8 -*-
"""生成姬小满校对稿 xlsx（D 列预填 ASR 原文，用户只需改有错的行）。"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ASR_LIST = Path(r"C:\Users\ASUS\Documents\AI换声\data\asr_out\姬小满_武道奇才\武道奇才 姬小满.list")
OUT = Path(r"C:\Users\ASUS\Documents\AI换声\data\asr_out\姬小满_校对稿_预填.xlsx")

BGM_FILES = {
    "姬小满-拾年声藏贺文.wav", "姬小满_武道奇才_回城1.wav", "姬小满_武道奇才_开场语音1.wav",
    "姬小满_武道奇才_挑衅1.wav", "姬小满_武道奇才_移动语音10.wav",
    "姬小满_武道奇才_购买装备-暗影战斧.wav",
}
SUSPECT = {
    "姬小满-拾年声藏贺文.wav": "长句疑似幻觉尾缀「我也来的愿望嘛」",
    "姬小满_武道奇才_大厅语音1.wav": "疑为「万法通变，我自成武道」类句式",
    "姬小满_武道奇才_大厅语音2.wav": "「强弱柔刚」疑为「刚柔并济」类成语",
    "姬小满_武道奇才_小满击杀，狂铁助攻2.wav": "「施工风景扯祸」明显不通",
    "姬小满_武道奇才_开场语音2.wav": "「蒸好」疑「正好」；「庄钟」疑「庄周」",
    "姬小满_武道奇才_开场语音3.wav": "与开场2内容应一致，请对照",
    "姬小满_武道奇才_技能1-1.wav": "疑为武术口令「起势。」",
    "姬小满_武道奇才_技能1-3.wav": "「鼓湖常试」不通",
    "姬小满_武道奇才_移动语音3.wav": "疑「万道长存」「小满独家」",
    "姬小满_武道奇才_移动语音7.wav": "确认是否哼唱",
    "姬小满_武道奇才_死亡语音1.wav": "「拜？」尾缀待确认",
    "姬小满_武道奇才_老夫子击杀，小满助攻.wav": "断句待确认",
    "姬小满_武道奇才_被动idle3.wav": "疑「我们走，皮皮虾」",
    "姬小满_武道奇才_被动idle5.wav": "疑「喝足吃饱」",
    "姬小满_武道奇才_移动语音11.wav": "疑「地下学院的人」",
    "姬小满_武道奇才_移动语音12.wav": "疑「只够装下身边的人」",
    "姬小满_武道奇才_复活语音1.wav": "「更像一条」疑误",
}

rows = []
with open(ASR_LIST, encoding="utf-8") as f:
    for line in f:
        line = line.rstrip("\n")
        if not line:
            continue
        parts = line.split("|")
        if len(parts) < 4:
            continue
        path, text = parts[0], "|".join(parts[3:])
        fname = Path(path).name
        marks, notes = [], []
        if fname in BGM_FILES:
            marks.append("BGM")
            notes.append("疑似带背景音乐，待定是否 UVR5")
        if fname in SUSPECT:
            marks.append("?")
            notes.append(SUSPECT[fname])
        rows.append([fname, text, "", " / ".join(marks) if marks else "", "；".join(notes)])

wb = Workbook()
ws = wb.active
ws.title = "校对稿"
headers = ["#", "文件名", "ASR自动识别（参考）", "校对后台词（请填，已预填ASR原文）", "标记", "备注"]
ws.append(headers)
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
    c.alignment = Alignment(horizontal="center", vertical="center")
for i, (fname, asr, _, marks, notes) in enumerate(rows, 1):
    ws.append([i, fname, asr, asr, marks, notes])
    r = ws.max_row
    if marks:
        for c in ws[r]:
            c.fill = PatternFill("solid", fgColor="FFF2CC")  # 浅黄高亮存疑行
    ws.cell(r, 4).font = Font(color="C00000")  # 校对列红色提醒可改

widths = [5, 42, 55, 55, 10, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = "A2"
wb.save(OUT)
print(f"已生成预填版校对稿：{OUT}")
print(f"共 {len(rows)} 条；存疑/BGM {sum(1 for r in rows if r[4])} 条已浅黄高亮")
