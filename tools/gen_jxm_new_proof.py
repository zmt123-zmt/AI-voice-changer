# -*- coding: utf-8 -*-
"""生成姬小满新皮肤（战舞者/灵喵仙官）校对稿 xlsx，D 列预填 ASR 原文。"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

DIRS = [
    (r"C:\Users\ASUS\Documents\AI换声\data\asr_out\姬小满_战舞者\战舞者 姬小满.list", "战舞者"),
    (r"C:\Users\ASUS\Documents\AI换声\data\asr_out\姬小满_灵喵仙官\灵喵仙官 姬小满.list", "灵喵仙官"),
]
OUT = Path(r"C:\Users\ASUS\Documents\AI换声\data\asr_out\姬小满_新皮肤_校对稿.xlsx")

# 依据 ASR 文本/王者语感标出的疑似错误（需听音频确认）
SUSPECT = {
    "战舞者 姬小满\\大厅语音1.wav": "「龟不哥贝儿哼人」疑为「乖不乖」类开场；整句待确认",
    "战舞者 姬小满\\大厅语音2.wav": "「姐不儿哼人」疑为「姐不乖」",
    "战舞者 姬小满\\大厅语音3.wav": "「龟不哥不儿哼人」疑为「乖不乖」类；「事先后退一步」疑为「先是后退一步」",
    "战舞者 姬小满\\技能1常规释放.wav": "「列书。」疑为技能口令（如「猎手/裂空」类）",
    "战舞者 姬小满\\技能2常规释放.wav": "待确认（技能口令短句）",
    "战舞者 姬小满\\技能3常规释放.wav": "待确认（技能口令短句）",
    "灵喵仙官 姬小满\\大厅喊话1.wav": "「缉拿腰挟黄蜂怪」疑为「缉拿妖邪黄蜂怪」；「硬蝉」疑为「硬茬」",
    "灵喵仙官 姬小满\\大厅喊话2.wav": "「遮妖路远不及一时长尺」疑为「遮妖路远，不急一时，长尺有度方至千里」",
    "灵喵仙官 姬小满\\击杀喊话1.wav": "「驱魔敕令风。」疑为「驱魔敕令，风！」类",
}

rows = []
for list_path, skin in DIRS:
    with open(list_path, encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            parts = line.split("|")
            if len(parts) < 4:
                continue
            path, text = parts[0], "|".join(parts[3:])
            fname = Path(path).name
            key = f"{skin}\\{fname}"
            marks = "?" if key in SUSPECT else ""
            notes = SUSPECT.get(key, "")
            rows.append([skin, fname, text, text, marks, notes])

wb = Workbook()
ws = wb.active
ws.title = "校对稿"
ws.append(["#", "皮肤", "文件名", "ASR自动识别（参考）", "校对后台词（请填，已预填ASR原文）", "标记", "备注"])
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
for i, (skin, fname, asr, _, marks, notes) in enumerate(rows, 1):
    ws.append([i, skin, fname, asr, asr, marks, notes])
    if marks:
        for c in ws[ws.max_row]:
            c.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(ws.max_row, 5).font = Font(color="C00000")
for col, w in zip("ABCDEFG", [5, 12, 38, 50, 50, 8, 35]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"
wb.save(OUT)
print(f"已生成: {OUT}（{len(rows)} 条，存疑 {len(SUSPECT)} 条已高亮）")
